import openpyxl
from datetime import date

# open spreadsheet with list of SKUs
# Read all SKUs and assign to a list item

skulist = []
skupath = "SKUs.xlsx"
skuwb = openpyxl.load_workbook(skupath)
skuws = skuwb.active
skurows = skuws.max_row
for i in range(1, skurows + 1):
    skulist.append(str(int(skuws.cell(row=i, column=1).value)))

# Open inventory spreadsheet
invpath = "inventory.xlsx"
invwb = openpyxl.load_workbook(invpath)
invws = invwb.active
invrows = invws.max_row

# Open cyclecount worksheet
countpath = "cyclecount.xlsx"
countwb = openpyxl.load_workbook(countpath)
countws = countwb.active

# write current date to cell B2 in the count sheet and set the starting row for the sku list
today = date.today()
countrow = 6

B2 = countws.cell(row=2, column=2)
B2.value = str(today)

# Foreach SKU in SKU list, if SKU matches value in column A:
    # write SKU to column B starting in row 6
    # write Description to column C
    # write bin location to column D
    # write quantity to column E
    # write comments to column G
for sku in skulist:
    for i in range(1, invrows + 1):
        cell_obj = str(invws.cell(row=i, column=1).value)
        if cell_obj == sku:
            # write SKU to column B starting in row 6
            countsku = countws.cell(row=countrow, column=2)
            countsku.value = cell_obj
            # write description to column C
            countdesc = countws.cell(row=countrow, column=3)
            countdesc.value = str(invws.cell(row=i, column=3).value)
            # write bin location to column D
            countbin = countws.cell(row=countrow, column=4)
            countbin.value = str(invws.cell(row=i, column=4).value)
            # write inventory quantity to column E
            countqty = countws.cell(row=countrow, column=5)
            countqty.value = str(invws.cell(row=i, column=5).value)
            # save the value of the comment/note to a variable
            # if the variable is not null, write the comment to column G
            comment = str(invws.cell(row=i, column=7).value)
            if comment != None:
                countcomm = countws.cell(row=countrow, column=7)
                countcomm.value = str(invws.cell(row=i, column=7).value)
            countrow = countrow + 1

# Save count workbook
countwb.save(countpath)
